##!/usr/bin/env python3
## -*- coding: utf-8 -*-
#"""
#Created on Fri Apr 24 00:38:11 2020
#
#@author: fabian
#"""
#
#import sys
#from PyQt5.QtGui import QImage, QPixmap
#from PyQt5.QtCore import QTimer
##from PyQt5.QtWidgets import QApplication, QDialog
#from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
#from PyQt5.uic import loadUi
#import cv2
#import numpy as np
#####mas claro para las imagenes
#from PyQt5 import QtCore, QtGui, QtWidgets,QtSql
##from verreg import *
#import cv2 
#import numpy as np
#font = cv2.FONT_HERSHEY_COMPLEX
#from PIL import Image
#import copy
#
#
#
#from ModuloUsuario import Usuario ###se importo el usuario y contraseña de la aplicacion 
#from tkinter import *
#from tkinter import messagebox
#import pymysql.cursors
#import mysql.connector
#from mysql.connector import Error
#from Modulo_Guardar import Guardando
#
#Error=Tk()   
##Error.eval('tk::PlaceWindow %s center ' % Error.winfo_toplevel())
#Error.withdraw()
#cedula=11
#
#Error=Tk()   
##Error.eval('tk::PlaceWindow %s center ' % Error.winfo_toplevel())
#Error.withdraw()
##n='f'
##
##if isinstance( n, ( int)):
##    
##    pass
##
##else:
##    print(" no entero")
##    QMessageBox.information( "GUARDAR","Datos guardados con éxito ")
#import sqlite3
#class Mostrar_Medir_Buscar(QMainWindow):
#    
#    def __init__(self,parent=None):
#        super(Mostrar_Medir_Buscar, self).__init__(parent)
#        loadUi('../Ventanas_Interfaz/Mostrar_Medir_Buscar.ui', self)
#        
#        self.boton_regresar.clicked.connect(self.abrirVentanaAnterior)
#        self.boton_mostrar.clicked.connect(self.mostar_imagen)
#        self.boton_guardar2.clicked.connect(self.guardar2)
#       # img=cv2.imread('pie1.png')
##        self.image_label3.setPixmap(img)
##        self.image_label3.setScaledContents(True)
#     
#    def mostar_imagen (self):
#       
#        self.image_label3.setPixmap(QtGui.QPixmap("Temporal_hsv.png"))
#        
#        
#    def abrirVentanaAnterior(self):
#        self.parent().show()
#        self.close() 
#        
#        
#    def guardar2(self):
#        global cedula
#        print(cedula)
#        # def guardar2(id3,foto_muestra,foto_medidas,foto):
#        foto_muestra=("../Imagenes_Huellas/Muestras_HSV/%d.png"%cedula)
#        foto_medidas=("../Imagenes_Huellas/Muestras_fotos/%d.png"%cedula)
#        foto=("../Imagenes_Huellas/Muestras_fotos/%d.png"%cedula)
#        Guardando.guardar2(self,cedula,foto_muestra,foto_muestra,foto_medidas,foto)
#        
#        
#        
#        
#if __name__ == '__main__':
#    app = QApplication(sys.argv)
#    window = Mostrar_Medir_Buscar()
#    #window.setWindowTitle('OpenCV Color Detector')
#    window.show()
#    sys.exit(app.exec_())
#
#        

#df1=()
#df=list(df1)
#for i in range (5):
#    df.append(i)
#print(df)
#
#df = pd.DataFrame([['a', 'b'], ['c', 'd']],
#                   index=['row 1', 'row 2'],
#                   columns=['col 1', 'col 2'])
#df.to_excel("output.xlsx")  

#datos=[]
#serie1 = pd.Series(datos)
#
#for i in range (5):
#    serie1 = serie1.append(pd.Series([i]))
#
#print("esta es la serie=\n",serie1)
#print("datos=\n",datos)


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Apr 20 16:47:58 2020

@author: fabian
"""

import sys
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer
#from PyQt5.QtWidgets import QApplication, QDialog
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5.uic import loadUi
import cv2
import numpy as np
####mas claro para las imagenes
from PyQt5 import QtCore, QtGui, QtWidgets,QtSql
#from verreg import *
import cv2 
import numpy as np
font = cv2.FONT_HERSHEY_COMPLEX
from PIL import Image
import copy


from PyQt5.QtCore import Qt


from tkinter import *
from tkinter import messagebox
import pymysql.cursors
import mysql.connector
from mysql.connector import Error
import pandas as pd
from openpyxl import Workbook
Error=Tk()   
#Error.eval('tk::PlaceWindow %s center ' % Error.winfo_toplevel())
Error.withdraw()
#n='f'
#
#if isinstance( n, ( int)):
#    
#    pass
#
#else:
#    print(" no entero")
#    QMessageBox.information( "GUARDAR","Datos guardados con éxito ")
import sqlite3
import xlsxwriter

#creamos el archivo xlsx y la hoja de trabajo
outWorkbook=xlsxwriter.Workbook('busquedad.xlsx')
outsheet=outWorkbook.add_worksheet()
class Buscar (QMainWindow):
            
    def __init__(self,parent=None):
        super(Buscar, self).__init__()
        loadUi('../Ventanas_Interfaz/Buscar.ui', self)
       
        self.boton_regresar.clicked.connect(self.regresar)
        self.Boton_Buscar.clicked.connect(self.Funcion_buscar)
        self.Boton_xlsx.clicked.connect(self.Funcion_Exportar_xlsx)
        
        #self.boton_regresar.clicked.connect(self.Funcion_buscar)
    def regresar():
        pass
        
        
    def Funcion_buscar(self):
        pass
        
    def Funcion_Exportar_xlsx(self):
        
     
        cc="*"
        
        try:
             rut=self.lineEdit_Ruta.text()             
             connection = mysql.connector.connect(host='localhost',                                            
                                                     database='base_huella',
                                                     user='root',
                                                     password='',
                                                     port='3311')
             if connection.is_connected():
                 db_Info = connection.get_server_info()
                 print(" Conectado al Servidor de MySQL version ", db_Info)
                 cursor = connection.cursor()
                 cursor.execute("select database();")
                 record = cursor.fetchone()
                 print("Ya esta conectado a la Base de Datos: ", record)
                
                 
 

                 

                                           
                                           
                 SQL= """SELECT %s FROM Registro   """%cc
                 cursor.execute(SQL)
                 results = cursor.fetchall()
                 wb = Workbook()
                 ws = wb.create_sheet(0)
                 ws.title = 'Resultado'
                 ws.append(cursor.column_names)
                 for row in results:
                     ws.append(row)
                
                 ruta = rut
                 wb.save(ruta + "/Base de datos JB-HUELLAS.xlsx")   

              
                 

#                 
                 
                 connection.commit()
                 print("Record inserted successfully into Laptop table")
                 QMessageBox.information(self, "Exportación","Exportación de datos con éxito ")
                 self.boton_regresar.setEnabled(True)
                 #self.thread.start()
#  
                 ######## fin de insertar los datos   #######
             else:
                 
                 QMessageBox.information(self, "verificar","Insertar la ruta donde desea guardar el documento ")
                
                 

        except mysql.connector.Error as error:
            connection.rollback()
            print("Error al conectar con MYSQL {}".format(error))
            QMessageBox.critical(self, "Error!","fallo al intentar conectar con MYSQL {}".format(error) )
        except PermissionError as error:
            print("Error en la exportación MYSQL {}".format(error))
            QMessageBox.critical(self, "Error!","fallo al intentar exportar XLSX {}".format(error) )
        
        
            
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Buscar()
    #window.setWindowTitle('OpenCV Color Detector')
    window.show()
    sys.exit(app.exec_())        
        
        




